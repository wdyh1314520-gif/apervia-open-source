# Auto-split helper: unified file diff router.
# Purpose: compare original/enhanced/diff candidates through one tool instead of
# ad-hoc shell commands or filename guessing.

from __future__ import annotations

import csv
import difflib
import io
import os
import re
import time
from dataclasses import dataclass, asdict

try:
    import openpyxl  # type: ignore
    from openpyxl import Workbook  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
except Exception:  # pragma: no cover - runtime may still have openpyxl through app imports
    openpyxl = None
    Workbook = None
    get_column_letter = None

FILE_DIFF_ROUTER_VERSION = 'file_diff_router_v2_generic_text_0606'

# Text/code file extensions that can be diffed as ordinary files.
# Keep this centralized so new formats are added once instead of scattered ifs.
_FDR_TEXT_DIFF_EXTS = {
    '.txt', '.md', '.markdown', '.csv', '.tsv', '.json', '.jsonl', '.yaml', '.yml', '.toml', '.ini', '.cfg', '.conf', '.env',
    '.py', '.js', '.mjs', '.cjs', '.ts', '.tsx', '.jsx', '.html', '.htm', '.css', '.scss', '.less', '.vue', '.svelte',
    '.xml', '.svg', '.sql', '.sh', '.bash', '.zsh', '.fish', '.ps1', '.bat', '.cmd',
    '.java', '.go', '.rs', '.c', '.h', '.cpp', '.cc', '.cxx', '.hpp', '.cs', '.php', '.rb', '.swift', '.kt', '.kts',
    '.dockerfile', '.gitignore', '.gitattributes', '.editorconfig', '.properties', '.gradle', '.lock', '.patch', '.diff', '.log'
}
_FDR_SPREADSHEET_DIFF_EXTS = {'.xlsx', '.xlsm'}
_FDR_BINARY_EXTS = {'.png', '.jpg', '.jpeg', '.webp', '.gif', '.bmp', '.ico', '.pdf', '.docx', '.pptx', '.zip', '.7z', '.rar', '.gz', '.tar', '.tgz', '.mp3', '.mp4', '.mov', '.avi'}
_FDR_TEXT_FILE_REPORT_RE = re.compile(r'(生成|导出|下载|保存|文件|补丁|patch|report|download|export|save)', re.I)


@dataclass(frozen=True)
class FileDiffPair:
    left_path: str
    right_path: str
    left_name: str
    right_name: str
    ext: str
    family: str

    def to_dict(self) -> dict:
        return asdict(self)


def _fdr_base(value: str = '') -> str:
    try:
        return os.path.basename(str(value or '').replace('\\', '/')).strip()
    except Exception:
        return str(value or '').strip()


def _fdr_ext(value: str = '') -> str:
    return os.path.splitext(_fdr_base(value))[1].lower()


def _fdr_text(value, limit=1000) -> str:
    s = '' if value is None else str(value)
    s = s.replace('\r\n', '\n').replace('\r', '\n')
    return s if len(s) <= limit else s[:limit] + '…'


def _fdr_resolve_file(path: str = '', messages=None) -> tuple[str, str]:
    resolver = globals().get('_sandbox_resolve_path')
    if callable(resolver):
        abs_path, rel = resolver(path, messages or [], must_exist=True, for_dir=False)
        return abs_path, rel
    p = str(path or '').strip()
    if not os.path.isabs(p):
        p = os.path.join('/mnt/data', p)
    if not os.path.isfile(p):
        raise FileNotFoundError(path)
    return p, os.path.relpath(p, '/mnt/data')


def _fdr_pick_pair(args: dict, messages=None) -> tuple[str, str, dict]:
    left = str(args.get('left_path') or args.get('old_path') or args.get('original_path') or args.get('before_path') or '').strip()
    right = str(args.get('right_path') or args.get('new_path') or args.get('enhanced_path') or args.get('after_path') or '').strip()
    ctx = {}
    if not (left and right):
        resolver = globals().get('file_context_resolve')
        if callable(resolver):
            ctx = resolver(query=str(args.get('query') or 'diff'), messages=messages or [], include_sandbox=True, max_files=160, max_candidates=8)
            candidates = [c for c in (ctx.get('compare_candidates') or []) if isinstance(c, dict)]
            if candidates:
                cand = candidates[0]
                l = cand.get('left') if isinstance(cand.get('left'), dict) else {}
                r = cand.get('right') if isinstance(cand.get('right'), dict) else {}
                left = left or str(l.get('path') or l.get('filename') or '').strip()
                right = right or str(r.get('path') or r.get('filename') or '').strip()
    if not (left and right):
        raise ValueError('missing_compare_pair')
    return left, right, ctx


def _fdr_lineage_records_for_pair(context: dict | None, *, left_rel: str = '', right_rel: str = '') -> tuple[dict, dict]:
    ctx = dict(context or {})
    cand = {}
    for item in (ctx.get('compare_candidates') or []):
        if not isinstance(item, dict):
            continue
        l = item.get('left') if isinstance(item.get('left'), dict) else {}
        r = item.get('right') if isinstance(item.get('right'), dict) else {}
        if (str(l.get('path') or l.get('filename') or '').strip() in {left_rel, _fdr_base(left_rel)} and
                str(r.get('path') or r.get('filename') or '').strip() in {right_rel, _fdr_base(right_rel)}):
            cand = item
            break
    if not cand and (ctx.get('compare_candidates') or []):
        first = ctx.get('compare_candidates')[0]
        cand = first if isinstance(first, dict) else {}
    left = cand.get('left') if isinstance(cand.get('left'), dict) else {}
    right = cand.get('right') if isinstance(cand.get('right'), dict) else {}
    maker = globals().get('file_lineage_make_record')
    if callable(maker):
        if not left:
            left = maker({'filename': _fdr_base(left_rel), 'path': left_rel, 'sandbox_path': left_rel, 'source': 'sandbox'}, source_hint='sandbox', idx=0) or {}
        if not right:
            right = maker({'filename': _fdr_base(right_rel), 'path': right_rel, 'sandbox_path': right_rel, 'source': 'sandbox'}, source_hint='sandbox', idx=1) or {}
    return dict(left or {}), dict(right or {})


def _fdr_make_output_audit(output_rel: str, output_abs: str, *, before_snapshot: dict, after_snapshot: dict, fmt: str, context: dict | None, left_rel: str, right_rel: str, family: str) -> dict:
    left_rec, right_rec = _fdr_lineage_records_for_pair(context, left_rel=left_rel, right_rel=right_rel)
    lineage_fn = globals().get('file_lineage_make_diff_lineage')
    lineage = {}
    if callable(lineage_fn):
        try:
            lineage = dict(lineage_fn(left=left_rec, right=right_rec, output_path=output_rel, output_filename=_fdr_base(output_rel), family=family))
        except Exception:
            lineage = {}
    if not lineage:
        lineage = {
            'role': 'diff',
            'lineage_key': str(family or '').strip(),
            'basis_filename': _fdr_base(left_rel),
            'source_paths': [left_rel, right_rel],
            'compared_paths': [left_rel, right_rel],
            'compared_filenames': [_fdr_base(left_rel), _fdr_base(right_rel)],
            'lineage_strength': 'strong',
        }
    audit_fn = globals().get('_sandbox_build_binary_audit')
    if callable(audit_fn):
        try:
            return audit_fn(output_rel, before_snapshot, after_snapshot, operation='sandbox_diff_files', fmt=fmt, lineage=lineage)
        except Exception:
            return {'_kind': 'file_edit_audit', 'operation': 'sandbox_diff_files', 'output_filename': output_rel, 'file_role': 'diff', **lineage}
    return {'_kind': 'file_edit_audit', 'operation': 'sandbox_diff_files', 'output_filename': output_rel, 'file_role': 'diff', **lineage}


def _fdr_cell_value(cell) -> str:
    v = getattr(cell, 'value', None)
    if v is None:
        return ''
    return str(v)


def _fdr_diff_xlsx(left_abs: str, right_abs: str, *, max_rows: int = 600, max_cols: int = 120, max_changes: int = 500) -> dict:
    if openpyxl is None:
        return {'ok': False, 'error': 'openpyxl_unavailable'}
    wb_l = openpyxl.load_workbook(left_abs, data_only=True, read_only=False)
    wb_r = openpyxl.load_workbook(right_abs, data_only=True, read_only=False)
    try:
        left_sheets = list(wb_l.sheetnames)
        right_sheets = list(wb_r.sheetnames)
        added_sheets = [s for s in right_sheets if s not in left_sheets]
        removed_sheets = [s for s in left_sheets if s not in right_sheets]
        common = [s for s in right_sheets if s in left_sheets]
        changes = []
        sheet_summaries = []
        total_changed = 0
        truncated = False
        for sheet in common:
            ws_l = wb_l[sheet]
            ws_r = wb_r[sheet]
            mr = min(max(int(ws_l.max_row or 0), int(ws_r.max_row or 0)), max_rows)
            mc = min(max(int(ws_l.max_column or 0), int(ws_r.max_column or 0)), max_cols)
            sheet_changed = 0
            for row in range(1, mr + 1):
                for col in range(1, mc + 1):
                    before = _fdr_cell_value(ws_l.cell(row=row, column=col))
                    after = _fdr_cell_value(ws_r.cell(row=row, column=col))
                    if before != after:
                        total_changed += 1
                        sheet_changed += 1
                        if len(changes) < max_changes:
                            col_letter = get_column_letter(col) if callable(get_column_letter) else str(col)
                            changes.append({
                                'sheet': sheet,
                                'cell': f'{col_letter}{row}',
                                'before': _fdr_text(before, 500),
                                'after': _fdr_text(after, 500),
                                'change_type': 'modified' if before and after else ('added' if after else 'removed'),
                            })
                        else:
                            truncated = True
            if int(ws_l.max_row or 0) != int(ws_r.max_row or 0) or int(ws_l.max_column or 0) != int(ws_r.max_column or 0) or sheet_changed:
                sheet_summaries.append({
                    'sheet': sheet,
                    'left_rows': int(ws_l.max_row or 0),
                    'right_rows': int(ws_r.max_row or 0),
                    'left_cols': int(ws_l.max_column or 0),
                    'right_cols': int(ws_r.max_column or 0),
                    'changed_cells': sheet_changed,
                })
        return {
            'ok': True,
            'type': 'xlsx_diff',
            'added_sheets': added_sheets,
            'removed_sheets': removed_sheets,
            'sheet_summaries': sheet_summaries,
            'changed_cell_count': total_changed,
            'changes': changes,
            'truncated': truncated,
        }
    finally:
        try:
            wb_l.close()
        except Exception:
            pass
        try:
            wb_r.close()
        except Exception:
            pass


def _fdr_is_probably_text(path: str, *, sample_bytes: int = 65536) -> tuple[bool, str]:
    """Return whether a file can be safely diffed as ordinary text/code.

    Extension is only a hint; the byte sample is the hard guard. This keeps the
    router generic without hard-coding every possible plain-text suffix.
    """
    ext = _fdr_ext(path)
    if ext in _FDR_BINARY_EXTS:
        return False, 'known_binary_extension'
    try:
        with open(path, 'rb') as f:
            data = f.read(max(1024, int(sample_bytes or 65536)))
    except Exception as exc:
        return False, f'read_sample_failed:{type(exc).__name__}'
    if b'\x00' in data:
        return False, 'contains_nul_byte'
    if not data:
        return True, 'empty_file'
    # Decode with replacement; too many replacement/control characters means binary-ish.
    text = data.decode('utf-8', errors='replace')
    repl = text.count('\ufffd')
    if repl > max(8, len(text) // 20):
        return False, 'too_many_decode_replacements'
    controls = sum(1 for ch in text if ord(ch) < 32 and ch not in '\n\r\t\f\b')
    if controls > max(16, len(text) // 25):
        return False, 'too_many_control_characters'
    return True, 'text_sample_ok'


def _fdr_read_text_lines(path: str, *, max_bytes: int = 4_000_000) -> tuple[list[str], bool]:
    size = 0
    try:
        size = int(os.path.getsize(path) or 0)
    except Exception:
        size = 0
    truncated = bool(size and size > max_bytes)
    with open(path, 'rb') as f:
        data = f.read(max_bytes + 1)
    if len(data) > max_bytes:
        data = data[:max_bytes]
        truncated = True
    text = data.decode('utf-8', errors='replace').replace('\r\n', '\n').replace('\r', '\n')
    return text.splitlines(), truncated


def _fdr_diff_text(left_abs: str, right_abs: str, *, max_lines: int = 300) -> dict:
    left_is_text, left_reason = _fdr_is_probably_text(left_abs)
    right_is_text, right_reason = _fdr_is_probably_text(right_abs)
    if not (left_is_text and right_is_text):
        return {
            'ok': False,
            'error': 'not_text_diffable',
            'left_reason': left_reason,
            'right_reason': right_reason,
            'supported_hint': 'Use explicit visual/binary handling for non-text files; ordinary diff supports text/code/config/data files.',
        }
    left_lines, left_truncated = _fdr_read_text_lines(left_abs)
    right_lines, right_truncated = _fdr_read_text_lines(right_abs)
    diff = list(difflib.unified_diff(left_lines, right_lines, fromfile=_fdr_base(left_abs), tofile=_fdr_base(right_abs), lineterm=''))
    add_count = sum(1 for line in diff if line.startswith('+') and not line.startswith('+++'))
    remove_count = sum(1 for line in diff if line.startswith('-') and not line.startswith('---'))
    return {
        'ok': True,
        'type': 'text_diff',
        'left_line_count': len(left_lines),
        'right_line_count': len(right_lines),
        'added_line_count': add_count,
        'removed_line_count': remove_count,
        'diff_line_count': len(diff),
        'diff_preview': diff[:max(20, min(max_lines, 1000))],
        'truncated': len(diff) > max_lines or left_truncated or right_truncated,
        'left_truncated': left_truncated,
        'right_truncated': right_truncated,
    }


def _fdr_write_text_report(diff: dict, output_abs: str, *, pair: FileDiffPair) -> int:
    lines = []
    lines.append(f'# File diff: {pair.left_name} -> {pair.right_name}')
    lines.append(f'# Type: {diff.get("type") or "text_diff"}')
    lines.append(f'# Added lines: {int(diff.get("added_line_count") or 0)}')
    lines.append(f'# Removed lines: {int(diff.get("removed_line_count") or 0)}')
    lines.append('')
    lines.extend(str(x) for x in (diff.get('diff_preview') or []))
    if diff.get('truncated'):
        lines.append('')
        lines.append('# NOTE: diff preview was truncated by max_changes/max_bytes limits.')
    os.makedirs(os.path.dirname(output_abs), exist_ok=True)
    with open(output_abs, 'w', encoding='utf-8', newline='\n') as f:
        f.write('\n'.join(lines) + '\n')
    return int(os.path.getsize(output_abs))


def _fdr_should_create_text_report(args: dict, query: str) -> bool:
    if 'create_file' in args:
        return bool(args.get('create_file'))
    return bool(_FDR_TEXT_FILE_REPORT_RE.search(str(query or '')))


def _fdr_write_xlsx_report(diff: dict, output_abs: str, *, pair: FileDiffPair) -> int:
    if Workbook is None:
        raise RuntimeError('openpyxl_unavailable')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Diff Summary'
    ws.append(['字段', '值'])
    ws.append(['原文件', pair.left_name])
    ws.append(['新版文件', pair.right_name])
    ws.append(['类型', diff.get('type')])
    ws.append(['新增 Sheet', ', '.join(diff.get('added_sheets') or [])])
    ws.append(['删除 Sheet', ', '.join(diff.get('removed_sheets') or [])])
    ws.append(['变化单元格数', int(diff.get('changed_cell_count') or 0)])
    ws.append(['是否截断', '是' if diff.get('truncated') else '否'])
    ws2 = wb.create_sheet('Sheet Summary')
    ws2.append(['Sheet', '原行数', '新版行数', '原列数', '新版列数', '变化单元格数'])
    for s in diff.get('sheet_summaries') or []:
        if isinstance(s, dict):
            ws2.append([s.get('sheet'), s.get('left_rows'), s.get('right_rows'), s.get('left_cols'), s.get('right_cols'), s.get('changed_cells')])
    ws3 = wb.create_sheet('Cell Changes')
    ws3.append(['Sheet', 'Cell', 'Change Type', 'Before', 'After'])
    for c in diff.get('changes') or []:
        if isinstance(c, dict):
            ws3.append([c.get('sheet'), c.get('cell'), c.get('change_type'), c.get('before'), c.get('after')])
    os.makedirs(os.path.dirname(output_abs), exist_ok=True)
    wb.save(output_abs)
    try:
        wb.close()
    except Exception:
        pass
    return int(os.path.getsize(output_abs))


def _file_diff_files_tool(args: dict | None = None, messages=None) -> dict:
    args = dict(args or {}) if isinstance(args, dict) else {}
    try:
        left_arg, right_arg, context = _fdr_pick_pair(args, messages=messages or [])
        left_abs, left_rel = _fdr_resolve_file(left_arg, messages=messages or [])
        right_abs, right_rel = _fdr_resolve_file(right_arg, messages=messages or [])
        ext = _fdr_ext(right_rel or right_abs) or _fdr_ext(left_rel or left_abs)
        family_fn = globals().get('file_context_normalized_family_name')
        family = family_fn(right_rel or left_rel) if callable(family_fn) else os.path.splitext(_fdr_base(right_rel or left_rel))[0]
        pair = FileDiffPair(left_path=left_rel, right_path=right_rel, left_name=_fdr_base(left_rel), right_name=_fdr_base(right_rel), ext=ext, family=str(family or '')).to_dict()
        max_changes = max(20, min(int(args.get('max_changes') or 500), 5000))
        if ext in _FDR_SPREADSHEET_DIFF_EXTS:
            diff = _fdr_diff_xlsx(left_abs, right_abs, max_changes=max_changes)
        else:
            left_textable, left_reason = _fdr_is_probably_text(left_abs)
            right_textable, right_reason = _fdr_is_probably_text(right_abs)
            if ext in _FDR_TEXT_DIFF_EXTS or (left_textable and right_textable):
                diff = _fdr_diff_text(left_abs, right_abs, max_lines=max_changes)
            else:
                diff = {'ok': False, 'error': 'unsupported_diff_format', 'supported': ['xlsx', 'xlsm', 'ordinary text/code/config/data files'], 'ext': ext, 'left_reason': left_reason, 'right_reason': right_reason}

        if not diff.get('ok'):
            return {'ok': False, '_kind': 'file_diff', 'version': FILE_DIFF_ROUTER_VERSION, 'pair': pair, **diff}
        query_text = str(args.get('query') or '')
        create_file = bool(args.get('create_file', ext in _FDR_SPREADSHEET_DIFF_EXTS)) if 'create_file' in args else (ext in _FDR_SPREADSHEET_DIFF_EXTS or _fdr_should_create_text_report(args, query_text))
        output_path = str(args.get('output_path') or '').strip()
        if create_file and ext in _FDR_SPREADSHEET_DIFF_EXTS:
            if not output_path:
                safe_family = re.sub(r'[\\/:*?"<>|]+', '_', str(family or os.path.splitext(_fdr_base(right_rel))[0])).strip(' _') or 'file'
                output_path = f'{safe_family}_diff对比表.xlsx'
            resolver = globals().get('_sandbox_resolve_path')
            if callable(resolver):
                output_abs, output_rel = resolver(output_path, messages or [], must_exist=False, for_dir=False)
            else:
                output_abs = output_path if os.path.isabs(output_path) else os.path.join('/mnt/data', output_path)
                output_rel = os.path.relpath(output_abs, '/mnt/data')
            before_snapshot = globals().get('_sandbox_file_binary_snapshot')(output_abs) if callable(globals().get('_sandbox_file_binary_snapshot')) else {'exists': os.path.exists(output_abs), 'size': os.path.getsize(output_abs) if os.path.exists(output_abs) else 0}
            size = _fdr_write_xlsx_report(diff, output_abs, pair=FileDiffPair(**pair))
            after_snapshot = globals().get('_sandbox_file_binary_snapshot')(output_abs) if callable(globals().get('_sandbox_file_binary_snapshot')) else {'exists': True, 'size': size}
            audit = _fdr_make_output_audit(output_rel, output_abs, before_snapshot=before_snapshot, after_snapshot=after_snapshot, fmt='xlsx', context=context, left_rel=left_rel, right_rel=right_rel, family=str(family or ''))
            diff['output_path'] = output_rel
            diff['output_size'] = size
            diff['file_edit_audit'] = audit
            diff['edit_audit'] = audit
            if bool(args.get('publish', True)):
                publish_fn = globals().get('_sandbox_publish_files_tool')
                if callable(publish_fn):
                    pub = publish_fn({'paths': [output_rel], 'force_zip': False, 'file_edit_audits': [audit]}, messages=messages or [])
                    diff['publish_result'] = pub if isinstance(pub, dict) else {}
                    if isinstance(pub, dict) and pub.get('ok'):
                        diff['files'] = pub.get('files') or []
                        diff['delivery_files'] = pub.get('delivery_files') or pub.get('files') or []
        elif create_file and diff.get('type') == 'text_diff':
            if not output_path:
                safe_family = re.sub(r'[\/:*?"<>|]+', '_', str(family or os.path.splitext(_fdr_base(right_rel))[0])).strip(' _') or 'file'
                output_path = f'{safe_family}.diff'
            resolver = globals().get('_sandbox_resolve_path')
            if callable(resolver):
                output_abs, output_rel = resolver(output_path, messages or [], must_exist=False, for_dir=False)
            else:
                output_abs = output_path if os.path.isabs(output_path) else os.path.join('/mnt/data', output_path)
                output_rel = os.path.relpath(output_abs, '/mnt/data')
            before_snapshot = globals().get('_sandbox_file_binary_snapshot')(output_abs) if callable(globals().get('_sandbox_file_binary_snapshot')) else {'exists': os.path.exists(output_abs), 'size': os.path.getsize(output_abs) if os.path.exists(output_abs) else 0}
            size = _fdr_write_text_report(diff, output_abs, pair=FileDiffPair(**pair))
            after_snapshot = globals().get('_sandbox_file_binary_snapshot')(output_abs) if callable(globals().get('_sandbox_file_binary_snapshot')) else {'exists': True, 'size': size}
            audit = _fdr_make_output_audit(output_rel, output_abs, before_snapshot=before_snapshot, after_snapshot=after_snapshot, fmt='diff', context=context, left_rel=left_rel, right_rel=right_rel, family=str(family or ''))
            diff['output_path'] = output_rel
            diff['output_size'] = size
            diff['file_edit_audit'] = audit
            diff['edit_audit'] = audit
            if bool(args.get('publish', True)):
                publish_fn = globals().get('_sandbox_publish_files_tool')
                if callable(publish_fn):
                    pub = publish_fn({'paths': [output_rel], 'force_zip': False, 'file_edit_audits': [audit]}, messages=messages or [])
                    diff['publish_result'] = pub if isinstance(pub, dict) else {}
                    if isinstance(pub, dict) and pub.get('ok'):
                        diff['files'] = pub.get('files') or []
                        diff['delivery_files'] = pub.get('delivery_files') or pub.get('files') or []
        summary_bits = []
        if diff.get('type') == 'xlsx_diff':
            summary_bits.append(f"变化单元格 {int(diff.get('changed_cell_count') or 0)} 个")
            if diff.get('added_sheets'):
                summary_bits.append('新增 sheet: ' + ', '.join(diff.get('added_sheets') or []))
            if diff.get('removed_sheets'):
                summary_bits.append('删除 sheet: ' + ', '.join(diff.get('removed_sheets') or []))
            if diff.get('truncated'):
                summary_bits.append('结果已截断')
        else:
            summary_bits.append(f"diff 行 {int(diff.get('diff_line_count') or 0)} 行")
            if diff.get('added_line_count') is not None or diff.get('removed_line_count') is not None:
                summary_bits.append(f"新增 {int(diff.get('added_line_count') or 0)} 行，删除 {int(diff.get('removed_line_count') or 0)} 行")
            if diff.get('truncated'):
                summary_bits.append('预览已截断')
        return {
            'ok': True,
            '_kind': 'file_diff',
            'version': FILE_DIFF_ROUTER_VERSION,
            'pair': pair,
            'summary': '；'.join(summary_bits),
            'diff': diff,
            'context_used': {'compare_candidates': (context or {}).get('compare_candidates', [])[:3]} if context else {},
            'fileNames': [pair.get('left_name'), pair.get('right_name')] + ([ _fdr_base(diff.get('output_path')) ] if diff.get('output_path') else []),
            'fileNameTotal': 3 if diff.get('output_path') else 2,
            'instruction': 'Use this diff result directly. If multiple compare_candidates existed and the chosen pair seems ambiguous, ask the user to choose another pair.',
        }
    except Exception as exc:
        return {'ok': False, '_kind': 'file_diff', 'version': FILE_DIFF_ROUTER_VERSION, 'error': f'{type(exc).__name__}: {exc}'}


def file_diff_policy_prompt() -> str:
    return (
        'FileDiffRouter 负责所有普通文件 diff/对比任务。先用 FileContextResolver 确定 original/enhanced/diff 候选；'
        'xlsx/xlsm 走表格 diff；普通文本/代码/配置/数据文件走统一文本 diff，可按需生成 .diff 补丁文件。多个候选不明确时让用户选择，不要靠 shell 列目录猜。'
    )
