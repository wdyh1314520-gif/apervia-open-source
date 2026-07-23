#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / '_shared'))
from sandbox_paths import resolve_mnt_path, rel_to_mnt  # noqa: E402


def main() -> int:
    parser = argparse.ArgumentParser(description='Probe XLSX workbook structure from /mnt/data.')
    parser.add_argument('path')
    parser.add_argument('--json', action='store_true')
    args = parser.parse_args()
    path = resolve_mnt_path(args.path)
    result = {'ok': False, 'path': rel_to_mnt(path), 'sheets': []}
    try:
        import openpyxl  # type: ignore
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=False)
        for ws in wb.worksheets:
            formulas = 0
            sampled = 0
            for row in ws.iter_rows(max_row=min(ws.max_row or 0, 200), max_col=min(ws.max_column or 0, 50)):
                for cell in row:
                    sampled += 1
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formulas += 1
            result['sheets'].append({'name': ws.title, 'max_row': ws.max_row, 'max_column': ws.max_column, 'sampled_cells': sampled, 'formula_cells_sampled': formulas})
        result['ok'] = True
    except Exception as exc:
        result['error'] = f'{type(exc).__name__}: {exc}'
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result.get('ok') else 2


if __name__ == '__main__':
    raise SystemExit(main())
